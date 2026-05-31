# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import io
import re
import os
import sys
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook


TEACHER_SHEET_CSV_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1lYaojsGUmca583pQ2MNnGPWd6UwzSl2511avnLqxRBM/export?format=csv&gid=0"
)
LIST_URL_TEMPLATE = "https://www.mimacstudy.com/tcher/studyQna/getStudyQnaList.ds?tcd={tcd}"
PROCESS_URL = "https://www.mimacstudy.com/tcher/studyQna/getStudyQnaProcess.ds"

ALLOWED_CATEGORIES = {"[강좌]", "[교재]", "[학습법]", "[기타]"}
SUBJECTS = [
    "한국지리", "세계지리", "동아시아사", "세계사",
    "생활과윤리", "윤리와사상", "사회문화", "정치와법",
    "경제", "통합사회", "한국사",
]
MAX_PAGES = 1000
MAX_WORKERS = 4
DATE_FORMAT = "%Y/%m/%d"
REQUEST_TIMEOUT = 30
REQUEST_RETRIES = 3


class DateRangeCancelled(Exception):
    pass


class CrawlerNetworkError(Exception):
    pass


def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


OUTPUT_DIR = app_dir() / "output"
LOG_PATH = OUTPUT_DIR / "crawler.log"


def show_message(title: str, message: str, *, is_error: bool = False) -> None:
    if os.environ.get("CRAWLER_NO_MESSAGE") == "1":
        return

    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        if is_error:
            messagebox.showerror(title, message)
        else:
            messagebox.showinfo(title, message)
        root.destroy()
    except Exception:
        pass


def normalize_teacher_code(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    code = str(value).strip()
    return code if code.isdigit() else None


def load_teachers() -> list[dict[str, str]]:
    try:
        response = requests.get(TEACHER_SHEET_CSV_URL, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
    except requests.exceptions.RequestException as error:
        raise CrawlerNetworkError(
            "구글 시트에서 선생님 목록을 불러오지 못했습니다. "
            "인터넷 연결과 구글 시트 공유 설정을 확인해 주세요."
        ) from error

    teachers: list[dict[str, str]] = []
    seen: set[str] = set()
    csv_text = response.content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(csv_text))

    for row in reader:
        if not row:
            continue
        code = normalize_teacher_code(row[0])
        if code and code not in seen:
            name = row[1].strip() if len(row) > 1 else ""
            teachers.append(
                {
                    "code": code,
                    "name": name or code,
                }
            )
            seen.add(code)

    if not teachers:
        raise ValueError("구글 시트의 A열에서 선생님 코드를 찾지 못했습니다.")

    return teachers


def prompt_crawl_options(teachers: list[dict[str, str]]) -> tuple[date, date, list[dict[str, str]]]:
    import tkinter as tk
    from tkinter import messagebox

    selected_options: dict[str, date | list[dict[str, str]]] = {}

    root = tk.Tk()
    root.title("크롤링 설정")
    root.resizable(False, False)

    frame = tk.Frame(root, padx=18, pady=16)
    frame.grid(row=0, column=0)

    tk.Label(frame, text="선생님 선택").grid(row=0, column=0, sticky="nw", pady=(0, 6))
    teacher_frame = tk.Frame(frame)
    teacher_frame.grid(
        row=0,
        column=1,
        sticky="w",
        padx=(10, 0),
        pady=(0, 10),
    )

    teacher_listbox = tk.Listbox(
        teacher_frame,
        width=28,
        height=min(max(len(teachers), 3), 10),
        selectmode=tk.EXTENDED,
        exportselection=False,
    )
    teacher_listbox.grid(row=0, column=0, sticky="nsew")

    teacher_scrollbar = tk.Scrollbar(
        teacher_frame,
        orient=tk.VERTICAL,
        command=teacher_listbox.yview,
    )
    teacher_scrollbar.grid(row=0, column=1, sticky="ns")
    teacher_listbox.configure(yscrollcommand=teacher_scrollbar.set)

    for teacher in teachers:
        teacher_listbox.insert(tk.END, f'{teacher["name"]} ({teacher["code"]})')

    def select_all_teachers() -> None:
        teacher_listbox.selection_set(0, tk.END)

    def clear_teacher_selection() -> None:
        teacher_listbox.selection_clear(0, tk.END)

    list_button_frame = tk.Frame(teacher_frame)
    list_button_frame.grid(row=1, column=0, columnspan=2, sticky="e", pady=(6, 0))
    tk.Button(list_button_frame, text="전체 선택", width=9, command=select_all_teachers).grid(row=0, column=0)
    tk.Button(list_button_frame, text="선택 해제", width=9, command=clear_teacher_selection).grid(
        row=0,
        column=1,
        padx=(6, 0),
    )

    tk.Label(frame, text="시작 날짜").grid(row=1, column=0, sticky="w", pady=(0, 6))
    start_entry = tk.Entry(frame, width=18)
    start_entry.grid(row=1, column=1, padx=(10, 0), pady=(0, 6))
    start_entry.insert(0, "2026/05/01")

    tk.Label(frame, text="종료 날짜").grid(row=2, column=0, sticky="w", pady=(0, 12))
    end_entry = tk.Entry(frame, width=18)
    end_entry.grid(row=2, column=1, padx=(10, 0), pady=(0, 12))
    end_entry.insert(0, "2026/05/11")

    tk.Label(frame, text="형식: YYYY/MM/DD").grid(
        row=3,
        column=0,
        columnspan=2,
        sticky="w",
        pady=(0, 12),
    )

    def submit() -> None:
        selected_indexes = teacher_listbox.curselection()
        if not selected_indexes:
            messagebox.showerror("입력 오류", "선생님을 한 명 이상 선택해 주세요.")
            return

        try:
            start_date = datetime.strptime(start_entry.get().strip(), DATE_FORMAT).date()
            end_date = datetime.strptime(end_entry.get().strip(), DATE_FORMAT).date()
        except ValueError:
            messagebox.showerror("입력 오류", "날짜는 YYYY/MM/DD 형식으로 입력해 주세요.")
            return

        if start_date > end_date:
            messagebox.showerror("입력 오류", "시작 날짜는 종료 날짜보다 늦을 수 없습니다.")
            return

        selected_options["start"] = start_date
        selected_options["end"] = end_date
        selected_options["teachers"] = [teachers[index] for index in selected_indexes]
        root.destroy()

    def cancel() -> None:
        root.destroy()

    button_frame = tk.Frame(frame)
    button_frame.grid(row=4, column=0, columnspan=2, sticky="e")
    tk.Button(button_frame, text="취소", width=8, command=cancel).grid(row=0, column=0)
    tk.Button(button_frame, text="시작", width=8, command=submit).grid(
        row=0,
        column=1,
        padx=(8, 0),
    )

    root.bind("<Return>", lambda _event: submit())
    root.bind("<Escape>", lambda _event: cancel())
    start_entry.focus_set()
    root.mainloop()

    if (
        "start" not in selected_options
        or "end" not in selected_options
        or "teachers" not in selected_options
    ):
        raise DateRangeCancelled

    return (
        selected_options["start"],
        selected_options["end"],
        selected_options["teachers"],
    )


def write_log(message: str) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with LOG_PATH.open("a", encoding="utf-8") as log_file:
        log_file.write(f"[{timestamp}] {message}\n")


def make_session(tcd: str) -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0 Safari/537.36"
            ),
            "Referer": LIST_URL_TEMPLATE.format(tcd=tcd),
        }
    )
    return session


def fetch_qna_page(session: requests.Session, page: int, tcd: str) -> str:
    payload = {
        "tcd": tcd,
        "srchWordType": "",
        "srchWordTxt": "",
        "pdsType": "",
        "pid": "",
        "ordType": "",
        "myQna": "N",
        "serverAddr": "",
        "banrDstin": "",
        "loDstin": "",
        "qnaStatus": "",
        "questType": "",
        "relm": "04",
        "tcdTabType": "tcdHome",
        "menuIdx": "",
        "relmName": "",
        "tcdName": "",
        "bNo": "",
        "currPage": str(page),
        "myQcurrPage": "",
        "isScrtN": "N",
        "procType": "",
        "ptype": "",
    }

    last_error: Exception | None = None
    for attempt in range(1, REQUEST_RETRIES + 1):
        try:
            response = session.post(
                PROCESS_URL,
                data=payload,
                headers={"X-Requested-With": "XMLHttpRequest"},
                timeout=REQUEST_TIMEOUT,
            )
            response.raise_for_status()
            return response.content.decode("euc-kr", errors="replace")
        except requests.exceptions.RequestException as error:
            last_error = error
            if attempt < REQUEST_RETRIES:
                time.sleep(attempt * 2)

    raise CrawlerNetworkError(
        "미맥스터디 서버에 연결하지 못했습니다. 인터넷 연결, VPN/방화벽, "
        "또는 사이트 접속 가능 여부를 확인한 뒤 다시 실행해 주세요."
    ) from last_error


def extract_category(title: str) -> str | None:
    match = re.match(r"^\[(강좌|교재|학습법|기타)\]", title)
    if not match:
        return None
    return f"[{match.group(1)}]"


def extract_course_name(summary: str) -> str:
    parts = [part.strip() for part in summary.split("·")]
    if len(parts) >= 3:
        return parts[-1]
    return summary.strip()


def parse_qna_items(html: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    items: list[dict[str, str]] = []

    for row in soup.select(".qnabbs_list li"):
        title_tag = row.select_one("a[href*='qnaDetail']")
        summary_tag = row.select_one("div.summary")
        date_tag = row.select_one("span.date")
        if title_tag is None or summary_tag is None or date_tag is None:
            continue

        title = title_tag.get_text(" ", strip=True)
        category = extract_category(title)
        if category not in ALLOWED_CATEGORIES:
            continue

        summary = " ".join(summary_tag.get_text(" ", strip=True).split())
        items.append(
            {
                "구분": category,
                "강좌명": extract_course_name(summary),
                "날짜": date_tag.get_text(strip=True),
            }
        )

    return items


def parse_page_dates(html: str) -> list:
    soup = BeautifulSoup(html, "html.parser")
    dates = []
    for date_tag in soup.select(".qnabbs_list li span.date"):
        date_text = date_tag.get_text(strip=True)
        dates.append(datetime.strptime(date_text, DATE_FORMAT).date())
    return dates


def crawl_items(start_date: date, end_date: date, tcd: str) -> list[dict[str, str]]:
    session = make_session(tcd)
    page_cache: dict[int, tuple[list[dict[str, str]], list]] = {}

    def get_page(page: int) -> tuple[list[dict[str, str]], list]:
        if page not in page_cache:
            html = fetch_qna_page(session, page, tcd)
            page_cache[page] = (parse_qna_items(html), parse_page_dates(html))
        return page_cache[page]

    def find_first_page(predicate) -> int | None:
        high = 1
        while high <= MAX_PAGES:
            _, dates = get_page(high)
            if not dates:
                return None
            if predicate(dates):
                break
            high *= 2
        else:
            return None

        low = (high // 2) + 1
        answer = high
        while low <= high:
            middle = (low + high) // 2
            _, dates = get_page(middle)
            if dates and predicate(dates):
                answer = middle
                high = middle - 1
            else:
                low = middle + 1
        return answer

    first_target_page = find_first_page(lambda dates: min(dates) <= end_date)
    if first_target_page is None:
        return []

    first_after_target_page = find_first_page(lambda dates: max(dates) < start_date)
    last_target_page = (
        first_after_target_page - 1 if first_after_target_page else MAX_PAGES
    )

    pages_to_fetch = range(first_target_page, last_target_page + 1)
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_page = {
            executor.submit(fetch_qna_page, make_session(tcd), page, tcd): page
            for page in pages_to_fetch
            if page not in page_cache
        }
        for future in as_completed(future_to_page):
            page = future_to_page[future]
            html = future.result()
            page_cache[page] = (parse_qna_items(html), parse_page_dates(html))

    matched_items: list[dict[str, str]] = []
    for page in pages_to_fetch:
        items, _ = get_page(page)
        for item in items:
            item_date = datetime.strptime(item["날짜"], DATE_FORMAT).date()
            if start_date <= item_date <= end_date:
                matched_items.append(item)

    return matched_items


def make_excel_bytes(items: list[dict[str, str]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "QNA"
    worksheet.append(["선생님 코드", "선생님 이름", "구분", "강좌명", "날짜"])

    for item in items:
        worksheet.append(
            [
                item["선생님 코드"],
                item["선생님 이름"],
                item["구분"],
                item["강좌명"],
                item["날짜"],
            ]
        )

    for column in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = min(
            max_length + 2, 80
        )

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def analyze1(items: list[dict[str, str]]) -> list[dict]:
    """날짜 × 과목 피벗 (분석1)"""
    from collections import defaultdict

    cols = SUBJECTS + ["그외"]
    data: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))

    for item in items:
        d = item["날짜"]
        lecture = item["강좌명"]
        matched = False
        for subj in SUBJECTS:
            if subj in lecture:
                data[d][subj] += 1
                matched = True
        if not matched:
            data[d]["그외"] += 1

    rows: list[dict] = []
    for d in sorted(data):
        row: dict = {"날짜": d}
        for col in cols:
            row[col] = data[d].get(col, 0)
        row["총 합계"] = sum(row[c] for c in cols)
        rows.append(row)

    if rows:
        total: dict = {"날짜": "합계"}
        for col in cols:
            total[col] = sum(r[col] for r in rows)
        total["총 합계"] = sum(r["총 합계"] for r in rows)
        rows.append(total)

    return rows


def analyze2(items: list[dict[str, str]]) -> list[dict]:
    """강좌명 × 구분 피벗 (분석2)"""
    from collections import defaultdict

    cats = ["[강좌]", "[교재]", "[학습법]", "[기타]"]
    data: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))

    for item in items:
        data[item["강좌명"]][item["구분"]] += 1

    rows: list[dict] = []
    for lecture, counts in data.items():
        row: dict = {"강좌명": lecture}
        for cat in cats:
            row[cat] = counts.get(cat, 0)
        row["총 합계"] = sum(row[c] for c in cats)
        rows.append(row)

    rows.sort(key=lambda r: r["총 합계"], reverse=True)

    if rows:
        total: dict = {"강좌명": "합계"}
        for cat in cats:
            total[cat] = sum(r[cat] for r in rows)
        total["총 합계"] = sum(r["총 합계"] for r in rows)
        rows.append(total)

    return rows


def _autofit_sheet(ws) -> None:
    for column in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in column), default=0)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 80)


def make_full_excel_bytes(
    items: list[dict[str, str]],
    rows1: list[dict],
    rows2: list[dict],
) -> bytes:
    from openpyxl.styles import Font

    workbook = Workbook()

    ws_qna = workbook.active
    ws_qna.title = "QNA"
    ws_qna.append(["선생님 코드", "선생님 이름", "구분", "강좌명", "날짜"])
    for item in items:
        ws_qna.append([item["선생님 코드"], item["선생님 이름"], item["구분"], item["강좌명"], item["날짜"]])
    _autofit_sheet(ws_qna)

    ws1 = workbook.create_sheet("분석1")
    if rows1:
        headers1 = list(rows1[0].keys())
        ws1.append(headers1)
        for row in rows1:
            ws1.append([row[h] for h in headers1])
        for cell in ws1[ws1.max_row]:
            cell.font = Font(bold=True)
        _autofit_sheet(ws1)

    ws2 = workbook.create_sheet("분석2")
    if rows2:
        headers2 = list(rows2[0].keys())
        ws2.append(headers2)
        for row in rows2:
            ws2.append([row[h] for h in headers2])
        for cell in ws2[ws2.max_row]:
            cell.font = Font(bold=True)
        _autofit_sheet(ws2)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def save_to_excel(items: list[dict[str, str]]) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"mimac_qna_{timestamp}.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "QNA"
    worksheet.append(["선생님 코드", "선생님 이름", "구분", "강좌명", "날짜"])

    for item in items:
        worksheet.append(
            [
                item["선생님 코드"],
                item["선생님 이름"],
                item["구분"],
                item["강좌명"],
                item["날짜"],
            ]
        )

    for column in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = min(
            max_length + 2,
            80,
        )

    workbook.save(output_path)
    return output_path


def run(start_date: date, end_date: date, teachers: list[dict[str, str]]) -> tuple[int, Path]:
    write_log("크롤링을 시작했습니다.")
    items: list[dict[str, str]] = []

    for teacher in teachers:
        tcd = teacher["code"]
        teacher_name = teacher["name"]
        write_log(f"{teacher_name}({tcd}) 크롤링을 시작했습니다.")
        tcd_items = crawl_items(start_date, end_date, tcd)
        for item in tcd_items:
            item["선생님 코드"] = tcd
            item["선생님 이름"] = teacher_name
        items.extend(tcd_items)
        write_log(f"{teacher_name}({tcd}): {len(tcd_items)}건")

    output_path = save_to_excel(items)
    write_log(f"{len(items)}건을 저장했습니다: {output_path}")
    return len(items), output_path


def main() -> None:
    try:
        teachers = load_teachers()
        start_date, end_date, selected_teachers = prompt_crawl_options(teachers)
        count, output_path = run(start_date, end_date, selected_teachers)
        message = (
            f"선택한 선생님 {len(selected_teachers)}명 기준으로 {count}건을 저장했습니다."
            f"\n\n저장 위치:\n{output_path}"
        )
        print(message)
        show_message("크롤링 완료", message)
    except DateRangeCancelled:
        print("크롤링 설정 입력이 취소되었습니다.")
    except CrawlerNetworkError as error:
        error_text = traceback.format_exc()
        write_log(error_text)
        show_message(
            "크롤링 오류",
            f"{error}\n\n로그 파일:\n{LOG_PATH}",
            is_error=True,
        )
    except Exception:
        error_text = traceback.format_exc()
        write_log(error_text)
        show_message(
            "크롤링 오류",
            f"오류가 발생했습니다.\n\n로그 파일을 확인해 주세요:\n{LOG_PATH}",
            is_error=True,
        )


if __name__ == "__main__":
    main()
